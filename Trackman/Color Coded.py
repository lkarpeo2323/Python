import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the CSV file
df = pd.read_csv('example_game_data.csv')

# Define the order of columns for optimized analysis
columns_to_keep = [
    # Pitch information
    'PitchNo',
    'PitchofPA',
    'PAofInning',
    'PitcherTeam',
    'Pitcher',
    'PitcherThrows',
    'BatterTeam',
    'Batter',
    'BatterSide',
    
    # Pitch characteristics
    'Pitch Velocity',
    'TaggedPitchType',
    'PitchCall',         
    'Hit Type',  # 'TaggedHitType' renamed to 'Hit Type'
    
    # Spin and angle data
    'VertRelAngle',
    'HorzRelAngle',
    'SpinRate',
    'SpinAxis',
    'Tilt',
    
    # Release and trajectory details
    'RelHeight',
    'RelSide',
    'Extension',
    'VertBreak',
    'HorzBreak',
    'PlateLocHeight',
    'PlateLocSide',
    
    # Flight and impact metrics
    'ZoneSpeed',
    'VertApprAngle',
    'HorzApprAngle',
    'ZoneTime',
    'Exit Velocity',  # 'ExitSpeed' was renamed to 'Exit Velocity'
    'Angle',
    'Direction',
    'Distance',
    
    # Results
    'PlayResult'
]

# Rename columns as needed
df.rename(columns={'RelSpeed': 'Pitch Velocity', 'ExitSpeed': 'Exit Velocity', 'TaggedHitType': 'Hit Type'}, inplace=True)

# Select the relevant columns in the specified order
cleaned_df = df[columns_to_keep]

# Save the cleaned data to a new Excel file
excel_path = 'cleaned_data.xlsx'
cleaned_df.to_excel(excel_path, index=False)

# Load the Excel file to apply color coding
wb = load_workbook(excel_path)
ws = wb.active

# Define color fills for important columns
important_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

# List of important columns to highlight
important_columns = ['Pitcher', 'Batter', 'Pitch Velocity', 'Exit Velocity']

# Apply the color coding to the entire columns of important columns
for col in important_columns:
    col_idx = columns_to_keep.index(col) + 1  # +1 because openpyxl is 1-indexed
    for row in range(1, ws.max_row + 1):  # Highlight the entire column
        ws.cell(row=row, column=col_idx).fill = important_fill

# Save the workbook with the applied styles
wb.save(excel_path)

# Print a message indicating completion
print(f"Cleaned and color-coded data has been saved to '{excel_path}'.")
